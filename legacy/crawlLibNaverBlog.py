import requests
import bs4
import json
import pymysql
import re
from datetime import datetime
from time import sleep
import openpyxl


class NaverBlogLister:

    def __init__(self, nUrl, keyword, channel, startDate, endDate):
        self.nUrl = nUrl
        self.keyword = keyword
        self.channel = channel
        self.startDate = startDate
        self.endDate = endDate

    def getNaverBlogUrl(self, idStr, noStr):
        urlStr = 'https://blog.naver.com/PostView.nhn?blogId=%s&logNo=%s&redirect=Dlog&widgetTypeCall=true&directAccess=false' % (
        idStr, noStr)
        return urlStr

    def createNaverBlogUrlList(self):
        wb = openpyxl.load_workbook('crawlstrength.xlsx')
        sheet1=wb['UrlLister']
        keyint = self.keyword.encode("UTF-8")
        keyhex = ""
        for i in keyint:
            keyhex = keyhex + '%{:02X}'.format(i)

        print(self.keyword, keyhex)

        # custom_header을 통해 아닌 것 처럼 위장하기
        custom_header = {
            'accept': 'application/json, text/plain, */*',
            'accept-encoding': 'gzip, deflate, br',
            'accept-language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
            'cookie': 'NNB=26JKENAB5VNFY; npic=LW1n/Y7lr1485twRJiDAu7IQNma7Byg1+eD1dujBnEarfw1Jj92XPcGCQrmTmIkDCA==; _ga=GA1.2.785938604.1550973896; nx_ssl=2; BMR=; _naver_usersession_=ldZTind9DyvToDWTHUQkNw==; page_uid=UfWJIwprvOsssOBmEzdssssstul-404192; JSESSIONID=FE270C9B4B35C84D83479B52E6020831.jvm1',
            'referer': 'https://section.blog.naver.com/Search/Post.nhn?pageNo=1&rangeType=ALL&orderBy=sim&keyword=' + keyhex,
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.100 Safari/537.36'}

        # 해당 접속 사이트가 아닌 원본데이터가 오는 url 추적. network에서 가지고 온다.

        pageNo = 1;
        nArticle = 0;
        counter = 0;
        er=0;
        exA= len(sheet1['A']) - 2
        exB=self.keyword
        exC=self.nUrl
        exD=self.channel
        exE=''
        exF=''
        exJ= 0;
        exK=''
        exH=''
        while True:
            url = "https://section.blog.naver.com/ajax/SearchList.nhn?countPerPage=7&currentPage=%d&endDate=%s" % (
            pageNo, self.endDate) \
                  + "&keyword=%s&orderBy=sim&startDate=%s&type=post" % (keyhex, self.startDate)

            req = requests.get(url, headers=custom_header)  # custom_header를 사용하지 않으면 접근 불가

            if req.status_code == requests.codes.ok:
                print("접속 성공")
                r = req.text[6:]

                data = json.loads(r)  # json에 반환된 데이터가 들어가 있다.
                print("Total count", data['result']['totalCount'])
                print(len(data['result']['searchList']), 'items found for '+ self.keyword)
####검색결과 없을 때 추가####
                if len(data['result']['searchList'])==0:
                    print("!!!!!!!no data to crawl,,ending process!!!!!!!")
                    self.nUrl = nArticle
                    exK='검색결과 존재하지 않음'
                    exG=''
                    exI= 0

                
                       
                conn = pymysql.connect(host='106.246.169.202', user='root', password='robot369',
                                       db='crawl', charset='utf8mb4')
#192.168.0.105
                curs = conn.cursor()

                for d in data['result']['searchList']:
                    idStr = d['blogId']
                    noStr = d['logNo']
                    pTimeStamp = d['addDate']
                    publishTime = datetime.fromtimestamp(pTimeStamp / 1e3)

                    
                   


                    
                    now = datetime.now()
                    dtStr = now.strftime("%Y-%m-%d %H:%M:%S")
                    print(dtStr)
                    print(d['blogId'], d['logNo'])
                    blogUrl = self.getNaverBlogUrl(idStr, noStr)
                    print(blogUrl)
                    sql0 = "select * from urllist where url=\'" + blogUrl + "\' " \
                           + "and keyword=\'"+self.keyword+"\' "\
                           + "and publishtime=\'%s\'"%(publishTime)
                    curs.execute(sql0)
                    rows = curs.fetchall()
                    print(rows,len(rows))
                    if len(rows)==0:
                        sql = "insert into urllist(keyword, channel, stdate, endate, accesstime, url, publishtime, crawled) " \
                              + "values (\'%s\', \'%s\', \'%s\', \'%s\', \'%s\',\'%s\',\'%s\' ,\'%s\')" % \
                              (self.keyword, self.channel, self.startDate, self.endDate, dtStr, blogUrl, publishTime, 'F')
                        curs.execute(sql)
                        conn.commit()
                    nArticle += 1
                    exI=nArticle
                    counter += 1
                    if (nArticle == 1):
                        for i in range(1,2,1):
                            exG=dtStr
                            print(exG)
                    if counter==3000:
                        sleep(120)
                        counter=0

                    if (nArticle >= self.nUrl):
                        break
#중복되는거 크롤링 안하게 세팅
                    
                        
            else:
                print("Error reading " + url)
                er += 1
                print("number of error:" + str(er))
                if er == 1:
                    now = datetime.now()
                    erdtStr = now.strftime("%Y-%m-%d %H:%M:%S")
                    exH = erdtStr
                    exJ = nArticle
                print("cannot access. waiting 4min for re-entry")
                sleep(240)

                    #sleep 추가하면 막힌거 다시 뚫을 수 있을수도 아니면 평균개수로 끊기..3000개로 안되는듯 , 엑셀 업데이트 되는거 다시확
            pageNo += 1

           
            if (nArticle >= self.nUrl):
                break
            #중복크롤링 방지
            if len(data['result']['searchList'])!=7:
                print("Process complete")
                break
            ##excel upload##
        now = datetime.now()
        endtStr = now.strftime("%Y-%m-%d %H:%M:%S")
        exL = endtStr
        sheet1.append([exA,exB,exC,exD,exE,exF,exG,exH,exI,exJ,exK,exL])
        wb.save('crawlstrength.xlsx')

        conn.close()




class NaverBlogCrawler:

    def __init__(self, keyword, channel, startDate, endDate):
        self.keyword = keyword
        self.channel = channel
        self.startDate = startDate
        self.endDate = endDate
        self.nCrawled = 0;
        self.exH =''
        self.er=0;
        self.exJ=0;
    def getHtmlSource(self, address):
        print('crawling ', address)
#####엑셀관련코드추가#####
        r = requests.get(address)
        
        
        if r.status_code == requests.codes.ok:
            print("접속 성공")
            t = r.text
            t = re.sub('<p', '\n<p', t, 0, re.I | re.S)
            t = re.sub('<br>', '\n<br>', t, 0, re.I | re.S)
        else:
            t = r.text
            print("Error while crawling " + address)
            self.er += 1
            print("number of error:" + str(self.er))
            if self.er == 1:
                now = datetime.now()
                erdtStr = now.strftime("%Y-%m-%d %H:%M:%S")
                self.exH = erdtStr
                self.exJ = self.nCrawled
            print("cannot access. waiting 4min for re-entry")
            sleep(240)
        return t

    def getPublishDate(self, soup):
        try:
            wpc = soup.find("span", {"class": "se_publishDate"})
            if wpc is None:
                return None
            str = wpc.get_text()
            print("Publish date "+str)
            str = re.sub(' ', '', str)
            l = re.split(':| |\.', str)
            now = datetime.now()
            dtStr = now.strftime("%Y-%m-%d %H:%M:%S")
            if len(l) == 3:
                dtStr = "%s-%s-%s %s:%s:%s" % (l[0], l[1], l[2], 0, 0, 0)
            elif len(l) == 5:
                dtStr = "%s-%s-%s %s:%s:%s" % (l[0], l[1], l[2], l[3], l[4], 0)
            else:
                return None

            return dtStr

        except Exception as ex:
            return None


    def getNReply(self, soup):
        try:
            wpc = soup.find("div", {"class": "wrap_postcomment"})
            cnt = int(wpc.find(text=re.compile('댓글')).next_sibling.get_text())
            print("got nReply-", cnt, end="\t")
            return cnt
        except Exception as ex:
            return 0


    def getNLike(self, soup):
        try:
            wpc = soup.find("div", {"class": "wrap_postcomment"})
            cnt = int(wpc.find("em", {"class": "u_cnt _count"}).get_text())
            print("got nLike-", cnt, end="\t")
            return cnt
        except Exception as ex:
            return 0


    def getIsCommercial(self, soup):
        try:
            comm = soup.find("img", {"src": re.compile('weble.net')})
            if comm:
                print("This is commerce", end="\t")
                return 'T'
            else:
                return 'F'
        except Exception as ex:
            return 'F'


    def extractBlogText(self, soup):

        bodySoup = soup.find_all('div', {'class': 'se-main-container'})

        print("LENGTH=%d"%(len(bodySoup)))
        
        if (len(bodySoup)==0):
            bodySoup = soup.find_all('div', {'class': 'se_paragraph'})

        if (len(bodySoup)==0):
            bodySoup = soup.find_all('div', {'id': 'postViewArea'})
            
        body = ''
        for bs in bodySoup:
            body = body + bs.getText()

        body = re.sub('<script.*?>.*?</script>', '', body, 0, re.I | re.S)
        body = re.sub('<style.*?>.*?</style>', '', body, 0, re.I | re.S)

        text = re.sub('<p', '\n<p', body, 0, re.I | re.S)
        text = re.sub('<br>', '\n', text, 0, re.I | re.S)
        text = re.sub('<.+?>', '', text, 0, re.I | re.S)

        text = re.sub(r'\\', '', text)
        text = re.sub('\'', '\\\'', text)
        text = re.sub('\"', '\\\"', text)

        text = re.sub('\u200b', ' ', text)
        text = re.sub('&nbsp;|\t', ' ', text)
        text = re.sub('\r\n', '\n', text)




        while (True):
            text = re.sub('  ', ' ', text)
            if text.count('  ') == 0:
                break

        while (True):
            text = re.sub('\n \n ', '\n', text)
            # print(text.count('\n \n '))
            if text.count('\n \n ') == 0:
                break

        while (True):
            text = re.sub(' \n', '\n', text)
            if text.count(' \n') == 0:
                break

        while (True):
            text = re.sub('\n ', '\n', text)
            if text.count('\n ') == 0:
                break

        while (True):
            text = re.sub('\n\n', '\n', text)
            # print(text.count('\n\n'))
            if text.count('\n\n') == 0:
                break

        # enclosed chars 
        text = re.sub(u'[\u2500-\u2BEF]', '', text) # I changed this to exclude chinese char

  
        #dingbats
        text = re.sub(u'[\u2702-\u27b0]', '', text)

        text = re.sub(u'[\uD800-\uDFFF]', '', text)
        text = re.sub(u'[\U0001F600-\U0001F64F]', '', text) # emoticons
        text = re.sub(u'[\U0001F300-\U0001F5FF]', '', text) # symbols & pictographs
        text = re.sub(u'[\U0001F680-\U0001F6FF]', '', text) # transport & map symbols
        text = re.sub(u'[\U0001F1E0-\U0001F1FF]', '', text) # flags (iOS)
        
        #text = re.sub(u'[\U0001F600-\U0001F1FF]', '', text)

        return text
    

    def crawlUrlTexts(self):
        self.nCrawled = 0
        wb = openpyxl.load_workbook('crawlstrength.xlsx')
        sheet1=wb['Crawler']

        conn = pymysql.connect(host='106.246.169.202', user='root', password='robot369',
                               db='crawl', charset='utf8mb4')
#106.246.169.202
        curs = conn.cursor(pymysql.cursors.DictCursor)
        sql = "select * from urllist where keyword=\'%s\' and channel=\'%s\' and publishtime>=\'%s\' and publishtime<=\'%s\' and crawled=\'%s\'" % \
              (self.keyword, self.channel, self.startDate, self.endDate, 'F')

        print(sql)
        curs.execute(sql)

        rows = curs.fetchall()
        print("Total %s urls targeted..."%(len(rows)))
        exA= len(sheet1['A']) - 2
        exB= self.keyword
        exC= len(rows)
        exD= self.channel
        exE=''
        exF=''
        exI=0
        exK='crawlUrlTexts'
        exL=''
                
        counter=0
        for row in rows:
            # print(row)
            

            blogUrl = row['url']
            pubTime = row['publishtime']
            print(blogUrl, pubTime)

            sql0="select count(*) from htdocs where url=\'" + blogUrl + "\'"\
                  + "and keyword=\'"+self.keyword+"\' "\
                  + "and publishtime=\'%s\'"%(pubTime)
 

            
            curs.execute(sql0)
            rows2 = curs.fetchall()
            print(rows2[0]['count(*)'])  
            if rows2[0]['count(*)'] == 0:
                now = datetime.now()
                dtStr = now.strftime("%Y-%m-%d %H:%M:%S")

                htmlSource = self.getHtmlSource(blogUrl)
                soup = bs4.BeautifulSoup(htmlSource, 'html.parser')

                #pubDate = self.getPublishDate(soup)

                nLike = self.getNLike(soup)
                nReply = self.getNReply(soup)
                isCommercial = self.getIsCommercial(soup)
                blogText = self.extractBlogText(soup)

                
                # print("pubdate = " +pubDate)
                if pubTime is not None:
                    sql = "insert into htdocs(keyword, stdate, endate, channel, url, accesstime, publishtime, nlike, nreply, iscommercial, htmltext)" \
                          + "values (\'%s\', \'%s\', \'%s\', \'%s\', \'%s\',\'%s\', \'%s\', \'%s\', \'%s\', \'%s\', \'%s\')" % \
                          (self.keyword, self.startDate, self.endDate, self.channel, blogUrl, dtStr, pubTime, nLike, nReply, isCommercial,
                           blogText)
                else:
                    sql = "insert into htdocs(keyword, stdate, endate, channel, url, accesstime, nlike, nreply, iscommercial, htmltext)" \
                          + "values (\'%s\', \'%s\', \'%s\', \'%s\', \'%s\',\'%s\', \'%s\', \'%s\', \'%s\', \'%s\')" % \
                          (self.keyword, self.startDate, self.endDate, self.channel, blogUrl, dtStr, nLike, nReply, isCommercial, blogText)

                curs.execute(sql)


            if pubTime is not None:
                sql = "update urllist set crawled='T' where url='%s'" % (blogUrl)\
                      + "and keyword=\'"+self.keyword+"\' "\
                      + "and publishtime=\'%s\'"%(pubTime)
            else:
                sql = "update urllist set crawled='T' where url='%s'" % (blogUrl)\
                      + "and keyword=\'"+self.keyword+"\' "\
                
 
            curs.execute(sql)
            
            conn.commit()


            self.nCrawled += 1
            print("Processed keyword %s, upto now %d article crawled" % (self.keyword, self.nCrawled))
            exI=self.nCrawled
            if (self.nCrawled == 1):
                for i in range(1,2,1):
                    exG=dtStr
                    print(exG)
                    

            counter += 1

            if counter==1000:
                sleep(30)
                counter=0
            
         ##excel upload##
        now = datetime.now()
        endtStr = now.strftime("%Y-%m-%d %H:%M:%S")
        exM = endtStr
        if exI != 0:
            sheet1.append([exA,exB,exC,exD,exE,exF,exG,self.exH,exI,self.exJ,exK,exL,exM])
            wb.save('crawlstrength.xlsx')



        conn.close()


    def crawlUrlTextsBatch(self):
        nCrawled = 0
        wb = openpyxl.load_workbook('crawlstrength.xlsx')
        sheet1=wb['Crawler']

        

        conn = pymysql.connect(host='106.246.169.202', user='root', password='robot369',
                               db='crawl', charset='utf8mb4')
#106.246.169.202
        curs = conn.cursor(pymysql.cursors.DictCursor)
        sql = "select * from urllist where crawled=\'F\' and channel = 'Naver Blog'"

        print(sql)
        curs.execute(sql)

        rows = curs.fetchall()
        print("Total %s urls targeted..."%(len(rows)))
        counter=0
        exI=0
        exA= len(sheet1['A']) - 2
        exB= '(batch)'
        exC= len((rows))
        exD= self.channel
        exE=''
        exF=''
        exK='crawlUrlTextsBatch'
        exL=''
        for row in rows:
            # print(row)
            keyword = row['keyword']
            channel = row['channel']
            blogUrl = row['url']
            pubTime = row['publishtime']
            startDate = row['stdate']
            endDate = row['endate']

            if pubTime is not None:
                sql0="select count(*) from htdocs where url=\'" + blogUrl + "\'"\
                  + "and keyword=\'"+keyword+"\' "\
                  + "and publishtime=\'%s\'"%(pubTime)
            else:
                sql0="select count(*) from htdocs where url=\'" + blogUrl + "\'"\
                  + "and keyword=\'"+keyword+"\'"
 

            
            curs.execute(sql0)
            rows2 = curs.fetchall()
            #print(rows2[0]['count(*)'])
            
            if rows2[0]['count(*)'] == 0:
                
                now = datetime.now()
                dtStr = now.strftime("%Y-%m-%d %H:%M:%S")

                htmlSource = self.getHtmlSource(blogUrl)
                soup = bs4.BeautifulSoup(htmlSource, 'html.parser')

                #pubDate = self.getPublishDate(soup)

                nLike = self.getNLike(soup)
                nReply = self.getNReply(soup)
                isCommercial = self.getIsCommercial(soup)
                blogText = self.extractBlogText(soup)

                
                # print("pubdate = " +pubDate)
                if pubTime is not None:
                    print(pubTime)
                    sql = "insert into htdocs(keyword, stdate, endate, channel, url, accesstime, publishtime, nlike, nreply, iscommercial, htmltext)" \
                          + "values (\'%s\', \'%s\', \'%s\', \'%s\', \'%s\',\'%s\', \'%s\', \'%s\', \'%s\', \'%s\', \'%s\')" % \
                          (keyword, startDate, endDate, channel, blogUrl, dtStr, pubTime, nLike, nReply, isCommercial,
                           blogText)
                else:
                    print("pubTime none")
                    sql = "insert into htdocs(keyword, stdate, endate, channel, url, accesstime, nlike, nreply, iscommercial, htmltext)" \
                          + "values (\'%s\', \'%s\', \'%s\', \'%s\', \'%s\',\'%s\', \'%s\', \'%s\', \'%s\', \'%s\')" % \
                          (keyword, startDate, endDate, channel, blogUrl, dtStr, nLike, nReply, isCommercial, blogText)

                curs.execute(sql)





            if pubTime is not None:
                sql = "update urllist set crawled='T' where url='%s'" % (blogUrl)\
                      + "and keyword=\'"+keyword+"\' "\
                      + "and publishtime=\'%s\'"%(pubTime)
            else:
                sql = "update urllist set crawled='T' where url='%s'" % (blogUrl)\
                      + "and keyword=\'"+keyword+"\' "\
                
 
            curs.execute(sql)
            
            conn.commit()


            nCrawled += 1
            print("Processed keyword %s, upto now %d article crawled" % (keyword, nCrawled))
            exI=nCrawled
            if (nCrawled == 1):
                for i in range(1,2,1):
                    exG=dtStr
                    print(exG)
            counter += 1
            if counter==1000:
                sleep(30)
                counter=0

                         
         ##excel upload##
        now = datetime.now()
        endtStr = now.strftime("%Y-%m-%d %H:%M:%S")
        exM = endtStr
        if exI != 0:
            sheet1.append([exA,exB,exC,exD,exE,exF,exG,self.exH,exI,self.exJ,exK,exL,exM])
            wb.save('crawlstrength.xlsx')




        conn.close()




